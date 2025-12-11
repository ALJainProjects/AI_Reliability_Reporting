"""Generate spreadsheet reports (CSV and Excel)."""

import csv
import logging
from datetime import datetime
from pathlib import Path

from ..models import Report

logger = logging.getLogger(__name__)


class SpreadsheetReporter:
    """Generate CSV and Excel spreadsheet reports."""

    def __init__(self):
        """Initialize the spreadsheet reporter."""
        pass

    def generate_csv(self, report: Report, output_path: Path | str) -> Path:
        """
        Generate a CSV file with incident data.

        Args:
            report: Report object with all data
            output_path: Path to save the CSV file

        Returns:
            Path to saved file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        headers = [
            "Date",
            "Time",
            "Company",
            "Title",
            "Impact",
            "Status",
            "Category",
            "Duration (hours)",
            "Summary",
            "Root Cause",
            "Affected Components",
            "Incident URL",
        ]

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            for incident in sorted(report.incidents, key=lambda x: x.created_at, reverse=True):
                # Get category name
                category = next(
                    (c for c in report.categories if c.id == incident.category),
                    None,
                )
                category_name = category.name if category else (incident.category or "Uncategorized")

                # Format duration
                duration = (
                    f"{incident.duration_hours:.2f}"
                    if incident.duration_hours
                    else ""
                )

                # Format components
                components = ", ".join(c.name for c in incident.affected_components)

                row = [
                    incident.created_at.strftime("%Y-%m-%d"),
                    incident.created_at.strftime("%H:%M:%S"),
                    incident.company_name,
                    incident.name,
                    incident.impact,
                    incident.status,
                    category_name,
                    duration,
                    incident.summary or "",
                    incident.root_cause or "",
                    components,
                    incident.shortlink or "",
                ]
                writer.writerow(row)

        logger.info(f"Saved CSV report to {output_path}")
        return output_path

    def generate_excel(self, report: Report, output_path: Path | str) -> Path:
        """
        Generate an Excel file with multiple sheets.

        Args:
            report: Report object with all data
            output_path: Path to save the Excel file

        Returns:
            Path to saved file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl required for Excel generation. Install with: pip install openpyxl"
            )

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Create sheets
        self._create_incidents_sheet(wb, report)
        self._create_categories_sheet(wb, report)
        self._create_statistics_sheet(wb, report)
        self._create_trends_sheet(wb, report)

        # Remove default sheet if it still exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(output_path)
        logger.info(f"Saved Excel report to {output_path}")
        return output_path

    def _create_incidents_sheet(self, wb, report: Report) -> None:
        """Create the incidents sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = wb.create_sheet("Incidents", 0)

        # Headers
        headers = [
            "Date",
            "Time",
            "Company",
            "Title",
            "Impact",
            "Status",
            "Category",
            "Duration (hours)",
            "Summary",
            "Root Cause",
            "Components",
            "URL",
        ]

        # Header styling
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Impact color mapping
        impact_colors = {
            "critical": PatternFill(
                start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"
            ),
            "major": PatternFill(
                start_color="FFA500", end_color="FFA500", fill_type="solid"
            ),
            "minor": PatternFill(
                start_color="FFD93D", end_color="FFD93D", fill_type="solid"
            ),
        }

        # Data rows
        for row_num, incident in enumerate(
            sorted(report.incidents, key=lambda x: x.created_at, reverse=True), 2
        ):
            category = next(
                (c for c in report.categories if c.id == incident.category), None
            )
            category_name = category.name if category else (incident.category or "Uncategorized")

            duration = incident.duration_hours if incident.duration_hours else None
            components = ", ".join(c.name for c in incident.affected_components)

            row_data = [
                incident.created_at.strftime("%Y-%m-%d"),
                incident.created_at.strftime("%H:%M:%S"),
                incident.company_name,
                incident.name,
                incident.impact,
                incident.status,
                category_name,
                duration,
                incident.summary or "",
                incident.root_cause or "",
                components,
                incident.shortlink or "",
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)

                # Apply impact coloring
                if col == 5 and value in impact_colors:  # Impact column
                    cell.fill = impact_colors[value]

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_categories_sheet(self, wb, report: Report) -> None:
        """Create the categories sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = wb.create_sheet("Categories")

        headers = ["Category ID", "Name", "Description", "Incident Count", "Keywords"]

        # Header styling
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_num, category in enumerate(
            sorted(report.categories, key=lambda x: -x.incident_count), 2
        ):
            row_data = [
                category.id,
                category.name,
                category.description,
                category.incident_count,
                ", ".join(category.keywords[:10]),
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)

        self._auto_adjust_columns(ws)

    def _create_statistics_sheet(self, wb, report: Report) -> None:
        """Create the statistics sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = wb.create_sheet("Statistics")

        # Header styling
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        label_font = Font(bold=True)

        stats = report.stats

        # Overview section
        ws.cell(row=1, column=1, value="Overview").font = header_font
        ws.merge_cells("A1:B1")
        ws["A1"].fill = header_fill

        overview_data = [
            ("Company", report.company_name),
            ("Analysis Period", f"{report.start_date.strftime('%Y-%m-%d')} to {report.end_date.strftime('%Y-%m-%d')}"),
            ("Total Days", report.timeframe_days),
            ("Total Incidents", stats.total_count),
            ("Resolved", stats.resolved_count),
            ("Unresolved", stats.unresolved_count),
        ]

        for row_num, (label, value) in enumerate(overview_data, 2):
            ws.cell(row=row_num, column=1, value=label).font = label_font
            ws.cell(row=row_num, column=2, value=value)

        # Impact section
        row_offset = len(overview_data) + 3
        ws.cell(row=row_offset, column=1, value="Impact Distribution").font = header_font
        ws.merge_cells(f"A{row_offset}:B{row_offset}")
        ws[f"A{row_offset}"].fill = header_fill

        impact_data = [
            ("Critical", stats.critical_count),
            ("Major", stats.major_count),
            ("Minor", stats.minor_count),
            ("None/Maintenance", stats.none_count),
        ]

        for row_num, (label, value) in enumerate(impact_data, row_offset + 1):
            ws.cell(row=row_num, column=1, value=label).font = label_font
            ws.cell(row=row_num, column=2, value=value)

        # Duration section
        row_offset = row_offset + len(impact_data) + 2
        ws.cell(row=row_offset, column=1, value="Duration Metrics (hours)").font = header_font
        ws.merge_cells(f"A{row_offset}:B{row_offset}")
        ws[f"A{row_offset}"].fill = header_fill

        duration_data = [
            ("MTTR", stats.mttr_hours),
            ("Average Duration", stats.avg_duration_hours),
            ("Median Duration", stats.median_duration_hours),
            ("Min Duration", stats.min_duration_hours),
            ("Max Duration", stats.max_duration_hours),
        ]

        for row_num, (label, value) in enumerate(duration_data, row_offset + 1):
            ws.cell(row=row_num, column=1, value=label).font = label_font
            ws.cell(row=row_num, column=2, value=round(value, 2) if value else "N/A")

        self._auto_adjust_columns(ws)

    def _create_trends_sheet(self, wb, report: Report) -> None:
        """Create the trends sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = wb.create_sheet("Trends")

        headers = [
            "Period",
            "Incidents",
            "Critical",
            "Major",
            "Avg Duration (hours)",
            "Total Downtime (hours)",
        ]

        # Header styling
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_num, trend in enumerate(report.trends, 2):
            row_data = [
                trend.period,
                trend.incident_count,
                trend.critical_count,
                trend.major_count,
                round(trend.avg_duration_hours, 2) if trend.avg_duration_hours else None,
                round(trend.total_downtime_hours, 2),
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)

        self._auto_adjust_columns(ws)

    def _auto_adjust_columns(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        from openpyxl.utils import get_column_letter

        for col in ws.columns:
            max_length = 0
            column = None

            for cell in col:
                try:
                    # Skip merged cells which don't have column_letter attribute
                    if hasattr(cell, 'column_letter'):
                        if column is None:
                            column = cell.column_letter
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = min(cell_length, 50)  # Cap at 50
                except Exception:
                    pass

            if column:
                adjusted_width = max(max_length + 2, 10)  # Minimum width of 10
                ws.column_dimensions[column].width = adjusted_width

    def save_all(
        self, report: Report, output_dir: Path | str, base_name: str | None = None
    ) -> tuple[Path, Path]:
        """
        Save both CSV and Excel reports.

        Args:
            report: Report object
            output_dir: Directory to save reports
            base_name: Base filename (without extension)

        Returns:
            Tuple of (csv_path, excel_path)
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        if base_name is None:
            date_str = datetime.now().strftime("%Y%m%d")
            base_name = f"{report.company_name.lower().replace(' ', '_')}_incidents_{date_str}"

        csv_path = self.generate_csv(report, output_dir / f"{base_name}.csv")
        excel_path = self.generate_excel(report, output_dir / f"{base_name}.xlsx")

        return csv_path, excel_path
